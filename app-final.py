from flask import Flask, request, jsonify, render_template, send_from_directory, session, redirect, url_for, send_file
from flask_cors import CORS
import sqlite3
import os
import json
from datetime import datetime
from PIL import Image
import pytesseract
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

app = Flask(__name__)
CORS(app)

# Configuración básica
app.secret_key = os.getenv('SECRET_KEY', 'gastos_app_secret_key_2025')
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'uploads')
DATABASE = os.getenv('DATABASE_URL', 'gastos.db').replace('sqlite:///', '') if os.getenv('DATABASE_URL', '').startswith('sqlite:///') else os.getenv('DATABASE_URL', 'gastos.db')

# Crear directorio de uploads
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Ruta principal - healthcheck
@app.route('/')
def index():
    if 'user' not in session:
        return redirect('/login')
    return render_template('index.html')

# Ruta de login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Autenticación simple
        if username == 'edurne' and password == 'edurne':
            session['user'] = {'username': 'edurne', 'role': 'admin'}
            return redirect('/')
        elif username == 'paul' and password == 'paul':
            session['user'] = {'username': 'paul', 'role': 'user'}
            return redirect('/')
        else:
            return render_template('login.html', error='Credenciales incorrectas')
    
    return render_template('login.html')

# Ruta de logout
@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect('/login')

# API para procesar imágenes con OCR
@app.route('/api/process-image', methods=['POST'])
def process_image():
    try:
        data = request.get_json()
        if not data or 'image' not in data:
            return jsonify({'error': 'No image provided'}), 400
        
        # Decodificar imagen base64
        image_data = data['image'].split(',')[1]  # Remover data:image/jpeg;base64,
        image_bytes = base64.b64decode(image_data)
        image = Image.open(io.BytesIO(image_bytes))
        
        # Mejorar imagen para OCR
        if image.mode != 'L':
            image = image.convert('L')
        
        # Mejorar contraste
        from PIL import ImageEnhance
        enhancer = ImageEnhance.Contrast(image)
        enhanced = enhancer.enhance(2.0)
        
        # Extraer texto con OCR
        text = pytesseract.image_to_string(enhanced, lang='spa+eng')
        
        # Procesar texto para extraer información
        lines = text.split('\n')
        amount = None
        date = None
        description = ""
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Buscar importe (patrones comunes)
            import re
            amount_match = re.search(r'(\d+[,.]?\d*)\s*€', line)
            if amount_match:
                amount = float(amount_match.group(1).replace(',', '.'))
            
            # Buscar fecha
            date_match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', line)
            if date_match:
                date = date_match.group(1)
            
            # Descripción (líneas más largas)
            if len(line) > 10 and not re.search(r'\d+[,.]?\d*', line):
                description += line + " "
        
        return jsonify({
            'success': True,
            'text': text,
            'amount': amount,
            'date': date,
            'description': description.strip()
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API para gastos
@app.route('/api/gastos', methods=['GET', 'POST'])
def gastos():
    if request.method == 'GET':
        try:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM gastos ORDER BY fecha DESC LIMIT 50')
            columns = [description[0] for description in cursor.description]
            gastos = [dict(zip(columns, row)) for row in cursor.fetchall()]
            conn.close()
            return jsonify(gastos)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO gastos (fecha, concepto, importe_eur, descripcion, imagen_path, usuario)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                data.get('fecha', datetime.now().strftime('%Y-%m-%d')),
                data.get('concepto', ''),
                data.get('importe_eur', 0),
                data.get('descripcion', ''),
                data.get('imagen_path', ''),
                session.get('user', {}).get('username', 'unknown')
            ))
            
            conn.commit()
            conn.close()
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

# API para exportar a Excel
@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    try:
        # Obtener parámetros de fecha
        fecha_inicio = request.args.get('fecha_inicio', '')
        fecha_fin = request.args.get('fecha_fin', '')
        
        # Obtener gastos
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        if fecha_inicio and fecha_fin:
            cursor.execute('''
                SELECT * FROM gastos 
                WHERE fecha BETWEEN ? AND ? 
                ORDER BY fecha DESC
            ''', (fecha_inicio, fecha_fin))
        else:
            cursor.execute('SELECT * FROM gastos ORDER BY fecha DESC')
        
        columns = [description[0] for description in cursor.description]
        gastos = [dict(zip(columns, row)) for row in cursor.fetchall()]
        conn.close()
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"
        
        # Headers
        headers = ['Fecha', 'Concepto', 'Importe EUR', 'Descripción', 'Usuario']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        for row_idx, gasto in enumerate(gastos, 2):
            ws.cell(row=row_idx, column=1, value=gasto.get('fecha', ''))
            ws.cell(row=row_idx, column=2, value=gasto.get('concepto', ''))
            ws.cell(row=row_idx, column=3, value=gasto.get('importe_eur', 0))
            ws.cell(row=row_idx, column=4, value=gasto.get('descripcion', ''))
            ws.cell(row=row_idx, column=5, value=gasto.get('usuario', ''))
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        return send_file(tmp_file_path, as_attachment=True, 
                        download_name=f'gastos_{fecha_inicio}_{fecha_fin}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API para exportar a PDF
@app.route('/api/export/pdf', methods=['GET'])
def export_pdf():
    try:
        # Obtener parámetros de fecha
        fecha_inicio = request.args.get('fecha_inicio', '')
        fecha_fin = request.args.get('fecha_fin', '')
        
        # Obtener gastos
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        if fecha_inicio and fecha_fin:
            cursor.execute('''
                SELECT * FROM gastos 
                WHERE fecha BETWEEN ? AND ? 
                ORDER BY fecha DESC
            ''', (fecha_inicio, fecha_fin))
        else:
            cursor.execute('SELECT * FROM gastos ORDER BY fecha DESC')
        
        columns = [description[0] for description in cursor.description]
        gastos = [dict(zip(columns, row)) for row in cursor.fetchall()]
        conn.close()
        
        # Crear PDF
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            doc = SimpleDocTemplate(tmp_file.name, pagesize=A4)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            # Título
            story.append(Paragraph("Reporte de Gastos", title_style))
            story.append(Spacer(1, 20))
            
            # Información del reporte
            info_text = f"Período: {fecha_inicio} - {fecha_fin}" if fecha_inicio and fecha_fin else "Todos los gastos"
            story.append(Paragraph(info_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Tabla de gastos
            if gastos:
                # Headers de la tabla
                table_data = [['Fecha', 'Concepto', 'Importe', 'Descripción', 'Usuario']]
                
                # Datos
                total = 0
                for gasto in gastos:
                    fecha = gasto.get('fecha', '')
                    concepto = gasto.get('concepto', '')
                    importe = gasto.get('importe_eur', 0)
                    descripcion = gasto.get('descripcion', '')
                    usuario = gasto.get('usuario', '')
                    
                    table_data.append([fecha, concepto, f"€{importe:.2f}", descripcion, usuario])
                    total += importe
                
                # Crear tabla
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(table)
                story.append(Spacer(1, 20))
                
                # Total
                total_style = ParagraphStyle(
                    'TotalStyle',
                    parent=styles['Heading2'],
                    fontSize=14,
                    textColor=colors.darkred
                )
                story.append(Paragraph(f"Total: €{total:.2f}", total_style))
            else:
                story.append(Paragraph("No hay gastos para el período seleccionado.", styles['Normal']))
            
            # Construir PDF
            doc.build(story)
            tmp_file_path = tmp_file.name
        
        return send_file(tmp_file_path, as_attachment=True, 
                        download_name=f'gastos_{fecha_inicio}_{fecha_fin}.pdf',
                        mimetype='application/pdf')
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Healthcheck específico
@app.route('/health')
def health():
    return jsonify({'status': 'healthy', 'message': 'GASTOS IA Final with OCR + Excel + PDF is running'})

if __name__ == '__main__':
    # Inicializar base de datos básica
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Crear tabla de gastos si no existe
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS gastos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                concepto TEXT,
                importe_eur REAL,
                descripcion TEXT,
                imagen_path TEXT,
                usuario TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Crear tabla de usuarios si no existe
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                name TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                parent_admin TEXT,
                active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Insertar usuarios por defecto si no existen
        cursor.execute('SELECT COUNT(*) FROM users')
        if cursor.fetchone()[0] == 0:
            cursor.execute('INSERT INTO users (username, password, name, role, parent_admin) VALUES (?, ?, ?, ?, ?)', 
                          ('edurne', 'edurne', 'Edurne', 'admin', None))
            cursor.execute('INSERT INTO users (username, password, name, role, parent_admin) VALUES (?, ?, ?, ?, ?)', 
                          ('paul', 'paul', 'Paul', 'user', 'edurne'))
        
        conn.commit()
        conn.close()
        print("✅ Base de datos inicializada correctamente")
    except Exception as e:
        print(f"❌ Error inicializando base de datos: {e}")
    
    # Iniciar aplicación
    port = int(os.getenv('PORT', 5100))
    print(f"🚀 Iniciando GASTOS IA Final con OCR + Excel + PDF en puerto {port}")
    app.run(debug=False, host='0.0.0.0', port=port)
